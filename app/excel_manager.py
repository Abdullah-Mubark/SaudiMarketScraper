import logging
import os
from datetime import datetime, date
from enum import Enum

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from analytics import Analytics
from config import config, Stock, StockData, FinancialIndicators
from scraper import FinancialIndicatorsColumnsIndex


class StocksTableColumns(Enum):
    SECTOR = 'Sector'
    STOCK_CODE = 'Stock Code'
    STOCK_NAME = 'Stock'
    PRICE = 'Price'
    COST_PRICE = 'Cost Price'
    PERCENTAGE_CHANGE_FROM_COST_PRICE = '% Diff from Cost Price'
    PERCENTAGE_OF_PORTFOLIO = '% of Portfolio'
    _52_WEEK_HIGH = '52 Week High'
    DIIFERENCE_FROM_52_WEEK_HIGH = 'Diff from 52 Week High'
    _52_WEEK_LOW = '52 Week Low'
    DIIFERENCE_FROM_52_WEEK_LOW = 'Diff from 52 Week Low'
    PRICE_EARNINGS_RATIO = 'P/E Ratio'
    FAIR_VALUE = 'Fair Value'
    DIFFERENCE_FROM_FAIR_VALUE = 'Diff from Fair Value'
    FAIR_VALUE_UNCERTAINTY = 'Fair Value Certainty'
    DIVIDEND_YIELD = '% Div Yield'


class ExcelManager:
    def __init__(self):
        self.analytics = Analytics()
        self.folder_name = config.excel.folder_name
        self.stocks_analysis_file_name = config.excel.stocks_analysis_file_name
        self.stocks_analysis_sheet_name = config.excel.stocks_analysis_sheet_name
        self.market_analysis_file_name = config.excel.market_analysis_file_name
        self.market_analysis_sheet_name = config.excel.market_analysis_sheet_name
        self.last_three_years = self.__get_last_three_years()
        self.stocks_table_name = config.excel.stocks_table_name
        self.stocks_table_start_row = config.excel.stocks_table_start_row
        self.market_table_start_row = config.excel.market_table_start_row
        self.stocks_table_cols = [col.value for col in StocksTableColumns]
        self.stocks_table_cols.extend(self.last_three_years)

        self.__create_folder_if_not_exists()

    def create_stocks_analysis(self, stock_data: StockData) -> None:
        if isinstance(stock_data.date, date):
            file_date = stock_data.date.strftime('%Y-%m-%d')
        else:
            file_date = stock_data.date

        file_name = f'{self.stocks_analysis_file_name}_{file_date}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = self.stocks_analysis_sheet_name

        self.__create_sheet_date(ws, stock_data.date)
        self.__create_stocks_table(ws, stock_data)

        wb.save(f'{self.folder_name}/{file_name}')
        logging.info(f'Stocks analysis excel file created: {self.folder_name}/{file_name}')

    def create_market_analysis(self, financial_indicators: FinancialIndicators) -> None:
        if isinstance(financial_indicators.date, date):
            file_date = financial_indicators.date.strftime('%Y-%m-%d')
        else:
            file_date = financial_indicators.date

        file_name = f'{self.market_analysis_file_name}_{file_date}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = self.market_analysis_sheet_name

        self.__create_sheet_date(ws, financial_indicators.date)
        self.__create_market_tables(ws, financial_indicators)

        wb.save(f'{self.folder_name}/{file_name}')
        logging.info(f'Market analysis excel file created: {self.folder_name}/{file_name}')

    def __create_folder_if_not_exists(self) -> None:
        if not os.path.exists(self.folder_name):
            os.makedirs(self.folder_name)

    def __get_last_three_years(self) -> list[str]:
        year = datetime.now().year
        return [str(year), str(year - 1), str(year - 2)]

    def __create_sheet_date(self, ws, date: date) -> None:
        ws['A1'] = 'Date'
        ws['B1'] = date

        ws['A1'].font = ws['B1'].font = Font(size=12, bold=True)

        ws['A1'].style = 'Headline 1'
        ws['B1'].style = 'Headline 2'

        ws['A1'].alignment = ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

        ws['B1'].number_format = 'dd/mm/yyyy'

    def __create_stocks_table(self, ws, stock_data: StockData) -> None:
        protfolio_total_value = self.analytics.get_portfolio_total_value(stock_data.stocks)

        ws.row_dimensions[self.stocks_table_start_row].height = 60
        for col_index, col in enumerate(self.stocks_table_cols):
            col_letter = f'{chr(65 + col_index)}'
            ws.column_dimensions[col_letter].width = 15

            cell_name = self.__get_cell_name_stock_table(col, self.stocks_table_start_row)
            ws[cell_name] = col
            ws[cell_name].font = Font(size=14, bold=True)
            ws[cell_name].style = 'Headline 1'
            ws[cell_name].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        dividend_cell_for_last_year = self.__get_cell_name_stock_table(self.last_three_years[0],
                                                                       self.stocks_table_start_row - 1)
        dividend_cell_for_three_years_ago = self.__get_cell_name_stock_table(self.last_three_years[2],
                                                                             self.stocks_table_start_row - 1)
        ws.merge_cells(f'{dividend_cell_for_last_year}:{dividend_cell_for_three_years_ago}')
        ws[dividend_cell_for_last_year].value = 'Dividends'
        ws[dividend_cell_for_last_year].font = Font(size=15, bold=True)
        ws[dividend_cell_for_last_year].style = 'Headline 1'
        ws[dividend_cell_for_last_year].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for index, stock in enumerate(stock_data.stocks):
            row_index = self.stocks_table_start_row + index + 1
            ws.row_dimensions[row_index].height = 30
            self.__create_stock_row(ws, stock, protfolio_total_value, row_index)

        table_start_cell = self.__get_cell_name_stock_table(self.stocks_table_cols[0], self.stocks_table_start_row)
        table_end_cell = self.__get_cell_name_stock_table(self.stocks_table_cols[-1],
                                                          self.stocks_table_start_row + len(stock_data.stocks) + 1)
        table = Table(displayName=self.stocks_table_name, ref=f"{table_start_cell}:{table_end_cell}")
        table.tableStyleInfo = TableStyleInfo(name='TableStyleLight20', showFirstColumn=False, showLastColumn=False,
                                              showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    def __create_stock_row(self, ws, stock: Stock, protfolio_total_value: float, row_index: int) -> None:
        sector_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.SECTOR.value, row_index)
        stock_code_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.STOCK_CODE.value, row_index)
        stock_name_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.STOCK_NAME.value, row_index)
        price_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.PRICE.value, row_index)
        cost_price_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.COST_PRICE.value, row_index)
        percentage_change_from_cost_price_cell_name = self.__get_cell_name_stock_table(
            StocksTableColumns.PERCENTAGE_CHANGE_FROM_COST_PRICE.value, row_index)
        price_earnings_ratio_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.PRICE_EARNINGS_RATIO.value,
                                                                          row_index)
        dividend_yield_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.DIVIDEND_YIELD.value, row_index)
        percentage_of_portfolio_cell_name = self.__get_cell_name_stock_table(
            StocksTableColumns.PERCENTAGE_OF_PORTFOLIO.value, row_index)
        _52_week_high_cell_name = self.__get_cell_name_stock_table(StocksTableColumns._52_WEEK_HIGH.value, row_index)
        difference_from_52_week_high_cell_name = self.__get_cell_name_stock_table(
            StocksTableColumns.DIIFERENCE_FROM_52_WEEK_HIGH.value, row_index)
        _52_week_low_cell_name = self.__get_cell_name_stock_table(StocksTableColumns._52_WEEK_LOW.value, row_index)
        difference_from_52_week_low_cell_name = self.__get_cell_name_stock_table(
            StocksTableColumns.DIIFERENCE_FROM_52_WEEK_LOW.value, row_index)
        fair_value_cell_name = self.__get_cell_name_stock_table(StocksTableColumns.FAIR_VALUE.value, row_index)
        difference_from_from_fair_value_cell_name = self.__get_cell_name_stock_table(
            StocksTableColumns.DIFFERENCE_FROM_FAIR_VALUE.value, row_index)
        fair_value_uncertainty_cell_name = self.__get_cell_name_stock_table(
            StocksTableColumns.FAIR_VALUE_UNCERTAINTY.value, row_index)
        last_year_dividend_cell_name = self.__get_cell_name_stock_table(self.last_three_years[0], row_index)
        two_years_ago_dividend_cell_name = self.__get_cell_name_stock_table(self.last_three_years[1], row_index)
        three_years_ago_dividend_cell_name = self.__get_cell_name_stock_table(self.last_three_years[2], row_index)

        ws[sector_cell_name] = '' if stock.sector is None else stock.sector
        ws[stock_code_cell_name] = '' if stock.code is None else stock.code
        ws[stock_name_cell_name] = '' if stock.name is None else stock.name

        ws[price_cell_name] = '' if stock.price is None else stock.price
        ws[cost_price_cell_name] = '' if stock.cost_price is None else stock.cost_price
        ws[
            percentage_change_from_cost_price_cell_name].value = f'=({price_cell_name}-{cost_price_cell_name})/{cost_price_cell_name}'
        ws[percentage_of_portfolio_cell_name] = '' if stock.price is None or stock.quantity_owned is None else (
                                                                                                                       stock.price * stock.quantity_owned) / protfolio_total_value

        ws[_52_week_high_cell_name] = '' if stock._52_week_high is None else stock._52_week_high
        ws[
            difference_from_52_week_high_cell_name].value = f'=({price_cell_name}-{_52_week_high_cell_name})/{_52_week_high_cell_name}'
        ws[_52_week_low_cell_name] = '' if stock._52_week_low is None else stock._52_week_low
        ws[
            difference_from_52_week_low_cell_name].value = f'=({price_cell_name}-{_52_week_low_cell_name})/{_52_week_low_cell_name}'

        ws[
            price_earnings_ratio_cell_name] = '' if stock.benchmark is None or stock.benchmark.p_e is None else stock.benchmark.p_e
        ws[
            dividend_yield_cell_name] = '' if stock.benchmark is None or stock.benchmark.div_yield is None else stock.benchmark.div_yield

        ws[
            fair_value_cell_name] = '' if stock.fair_value is None or stock.fair_value.average is None else stock.fair_value.average
        ws[
            difference_from_from_fair_value_cell_name].value = f'=({price_cell_name}-{fair_value_cell_name})/{fair_value_cell_name}'
        ws[
            fair_value_uncertainty_cell_name] = '' if stock.fair_value is None or stock.fair_value.uncertainty is None else stock.fair_value.uncertainty

        ws[last_year_dividend_cell_name] = self.analytics.get_dividends_sum_amount_for_year(stock, int(
            self.last_three_years[0]))
        ws[two_years_ago_dividend_cell_name] = self.analytics.get_dividends_sum_amount_for_year(stock, int(
            self.last_three_years[1]))
        ws[three_years_ago_dividend_cell_name] = self.analytics.get_dividends_sum_amount_for_year(stock, int(
            self.last_three_years[2]))

        self.__center_align_cells(ws, [stock_code_cell_name, price_cell_name, cost_price_cell_name,
                                       percentage_change_from_cost_price_cell_name,
                                       dividend_yield_cell_name, last_year_dividend_cell_name,
                                       two_years_ago_dividend_cell_name, three_years_ago_dividend_cell_name,
                                       fair_value_uncertainty_cell_name,
                                       percentage_of_portfolio_cell_name, _52_week_high_cell_name,
                                       difference_from_52_week_high_cell_name, _52_week_low_cell_name,
                                       difference_from_52_week_low_cell_name, fair_value_cell_name,
                                       difference_from_from_fair_value_cell_name, price_earnings_ratio_cell_name])

        self.__center_align_cells(ws, [sector_cell_name, stock_name_cell_name],
                                  wrap_text=True)

        self.__set_cell_number_format(ws, [price_cell_name, cost_price_cell_name, price_earnings_ratio_cell_name,
                                           _52_week_high_cell_name, _52_week_low_cell_name,
                                           fair_value_cell_name, last_year_dividend_cell_name,
                                           two_years_ago_dividend_cell_name, three_years_ago_dividend_cell_name],
                                      '0.00')

        self.__set_cell_number_format(ws, [percentage_of_portfolio_cell_name, dividend_yield_cell_name,
                                           percentage_change_from_cost_price_cell_name,
                                           difference_from_from_fair_value_cell_name,
                                           difference_from_52_week_high_cell_name,
                                           difference_from_52_week_low_cell_name], '0.00%')

        self.__bold_cells(ws,
                          [price_cell_name, percentage_change_from_cost_price_cell_name, price_earnings_ratio_cell_name,
                           dividend_yield_cell_name, _52_week_high_cell_name, _52_week_low_cell_name])

        ws.conditional_formatting.add(percentage_change_from_cost_price_cell_name,
                                      ColorScaleRule(start_type='num', start_value=0, start_color='c6efce',
                                                     end_type='num', end_value=0, end_color='ffc7ce'))
        ws.conditional_formatting.add(difference_from_from_fair_value_cell_name,
                                      ColorScaleRule(start_type='num', start_value=0, start_color='c6efce',
                                                     end_type='num', end_value=0, end_color='ffc7ce'))

    def __create_market_tables(self, ws, financial_indicators: FinancialIndicators):
        table_start_row = self.market_table_start_row

        for col_index, _ in enumerate(financial_indicators.headers):
            col_letter = f'{chr(65 + col_index)}'
            ws.column_dimensions[col_letter].width = 15

        for industry_group in financial_indicators.industry_groups:
            industry_group_stocks = [stock for stock in financial_indicators.stocks if
                                     stock.industry_group == industry_group]

            if len(industry_group_stocks) == 0:
                continue

            self.__create_market_table(ws, industry_group_stocks, financial_indicators.headers, table_start_row,
                                       industry_group)
            table_start_row += len(industry_group_stocks) + 4

    def __create_market_table(self, ws, stocks: list[Stock], headers: list[str], start_row: int, industry_group: str):
        ws.row_dimensions[start_row].height = 60
        for col_index, col in enumerate(headers):
            cell_name = self.__get_cell_name(col_index, start_row)
            ws[cell_name] = col
            ws[cell_name].font = Font(size=14, bold=True)
            ws[cell_name].style = 'Headline 1'
            ws[cell_name].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        industry_group_col_first_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.INDUSTRY_GROUP.value,
                                                             start_row + 1)
        industry_group_col_last_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.INDUSTRY_GROUP.value,
                                                            len(stocks) + start_row)
        ws.merge_cells(f'{industry_group_col_first_cell}:{industry_group_col_last_cell}')
        ws[industry_group_col_first_cell].value = industry_group
        ws[industry_group_col_first_cell].font = Font(size=15, bold=True)
        ws[industry_group_col_first_cell].style = 'Headline 1'
        ws[industry_group_col_first_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        table_row = start_row + 1
        for stock in stocks:
            ws.row_dimensions[table_row].height = 30
            self.__create_market_table_row(ws, stock, table_row)
            table_row += 1

        table_start_cell = self.__get_cell_name(1, start_row)
        table_end_cell = self.__get_cell_name(len(headers) - 1, len(stocks) + start_row)
        table = Table(displayName=industry_group.replace(" ", ""), ref=f"{table_start_cell}:{table_end_cell}")
        table.tableStyleInfo = TableStyleInfo(name='TableStyleLight20', showFirstColumn=False, showLastColumn=False,
                                              showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    def __create_market_table_row(self, ws, stock: Stock, start_row: int):
        stock_name_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.COMPANY.value + 1, start_row)
        stock_price_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.PRICE.value + 1, start_row)
        stock_issued_shares_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.ISSUED_SHARES.value + 1,
                                                        start_row)
        stock_net_income_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.NET_INCOME.value + 1, start_row)
        stock_shareholders_equity_cell = self.__get_cell_name(
            FinancialIndicatorsColumnsIndex.SHAREHOLDERS_EQUITY.value + 1, start_row)
        stock_market_cap_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.MARKET_CAP.value + 1, start_row)
        stock_market_cap_percentage_cell = self.__get_cell_name(
            FinancialIndicatorsColumnsIndex.MARKET_CAP_PERCENTAGE.value + 1, start_row)
        stock_earnings_per_share_cell = self.__get_cell_name(
            FinancialIndicatorsColumnsIndex.EARNINGS_PER_SHARE.value + 1, start_row)
        stock_pe_ratio_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.P_E_RATIO.value + 1, start_row)
        stock_book_value_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.BOOK_VALUE.value + 1, start_row)
        stock_pb_ratio_cell = self.__get_cell_name(FinancialIndicatorsColumnsIndex.P_B_RATIO.value + 1, start_row)

        ws[stock_name_cell] = stock.name if stock.name else ''
        ws[stock_price_cell] = stock.price if stock.price else ''
        ws[stock_issued_shares_cell] = stock.issued_shares if stock.issued_shares else ''
        ws[stock_net_income_cell] = stock.net_profit if stock.net_profit else ''
        ws[stock_shareholders_equity_cell] = stock.shareholders_equity if stock.shareholders_equity else ''
        ws[stock_market_cap_cell] = stock.market_cap if stock.market_cap else ''
        ws[stock_market_cap_percentage_cell] = stock.market_cap_percentage if stock.market_cap_percentage else ''
        ws[stock_earnings_per_share_cell] = stock.earnings_per_share if stock.earnings_per_share else ''
        ws[stock_pe_ratio_cell] = stock.benchmark.p_e if stock.benchmark.p_e else ''
        ws[stock_book_value_cell] = stock.book_value_per_share if stock.book_value_per_share else ''
        ws[stock_pb_ratio_cell] = stock.benchmark.p_b if stock.benchmark.p_b else ''

        self.__center_align_cells(ws, [stock_price_cell, stock_issued_shares_cell, stock_net_income_cell,
                                       stock_shareholders_equity_cell, stock_market_cap_cell,
                                       stock_market_cap_percentage_cell, stock_earnings_per_share_cell,
                                       stock_pe_ratio_cell, stock_book_value_cell, stock_pb_ratio_cell])

        self.__center_align_cells(ws, [stock_name_cell], wrap_text=True)

        self.__bold_cells(ws, [stock_net_income_cell, stock_market_cap_cell, stock_pe_ratio_cell, stock_pb_ratio_cell])

        self.__set_cell_number_format(ws, [stock_price_cell, stock_issued_shares_cell, stock_net_income_cell,
                                           stock_shareholders_equity_cell, stock_market_cap_cell,
                                           stock_earnings_per_share_cell, stock_pe_ratio_cell, stock_book_value_cell,
                                           stock_pb_ratio_cell], '0.00')

        self.__set_cell_number_format(ws, [stock_market_cap_percentage_cell], '0.00%')

    def __get_cell_name_stock_table(self, col_name: str, row_index: int) -> str:
        col_index = self.stocks_table_cols.index(col_name)
        return f'{chr(65 + col_index)}{row_index}'

    def __get_cell_name(self, col_index: int, row_index: int) -> str:
        return f'{chr(65 + col_index)}{row_index}'

    def __center_align_cells(self, ws, cells: str, wrap_text=False) -> None:
        for cell in cells:
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap_text)

    def __bold_cells(self, ws, cells: str) -> None:
        for cell in cells:
            ws[cell].font = Font(bold=True)

    def __set_cell_number_format(self, ws, cells: str, format: str) -> None:
        for cell in cells:
            ws[cell].number_format = format
